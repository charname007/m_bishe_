"""
微博Excel数据导入PostgreSQL脚本
从 data/corpuses/*.xlsx 导入微博数据
只保留通用字段，适配知识图谱抽取
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import execute_batch
from datetime import datetime
import json
import re
import hashlib
from glob import glob

from utils.text_preprocessor import WeiboTextPreprocessor

# 从.env加载配置
from dotenv import load_dotenv
load_dotenv()

DB_CONFIG = {
    'host': os.getenv('PG_HOST', 'localhost'),
    'port': int(os.getenv('PG_PORT', 5432)),
    'database': os.getenv('PG_DATABASE', 'bishe'),
    'user': os.getenv('PG_USER', 'postgres'),
    'password': os.getenv('PG_PASSWORD', 'postgres')
}

# Excel文件目录
EXCEL_DIR = 'data/corpuses'
# 测试模式：只处理指定文件（为空则处理全部）
TEST_FILE = None  # '微博搜索关键词采集(4).xlsx' 小文件测试时设置

# 通用表名（支持微博和小红书）
TABLE_NAME = 'social_media_notes'

# 预处理器
PREPROCESSOR = WeiboTextPreprocessor(convert_to_simplified=True)


def generate_id(row: dict) -> str:
    """生成唯一ID（基于详情链接或内容hash）"""
    if row.get('详情链接'):
        return hashlib.md5(row['详情链接'].encode()).hexdigest()[:16]
    content = row.get('博文内容', '') or ''
    return hashlib.md5(f"{row.get('博主昵称', '')}{content}".encode()).hexdigest()[:16]


def parse_weibo_time(time_str: str) -> datetime:
    """解析微博时间格式: 2024年06月11日 15:19"""
    if not time_str:
        return None
    try:
        match = re.match(r'(\d{4})年(\d{2})月(\d{2})日\s+(\d{2}):(\d{2})', time_str)
        if match:
            return datetime(
                int(match.group(1)),
                int(match.group(2)),
                int(match.group(3)),
                int(match.group(4)),
                int(match.group(5))
            )
    except (ValueError, TypeError):
        pass
    return None


def parse_int_safe(value) -> int:
    """安全转整数"""
    if value is None or value == '':
        return 0
    try:
        return int(float(str(value).strip()))
    except:
        return 0


def preprocess_weibo_note(row: dict) -> dict:
    """预处理微博数据"""
    # 处理博文内容
    content_result = PREPROCESSOR.preprocess(row.get('博文内容', ''))

    # 短文本过滤检查
    if content_result.get('should_filter'):
        return None  # 返回None表示需要过滤

    # 提取微博话题 #xxx#
    topics = content_result['topics']
    # 补充关键词作为话题
    keyword = row.get('关键词', '')
    if keyword and keyword not in topics:
        keyword_clean = re.sub(r'^#|#$', '', keyword)
        topics.append(keyword_clean)

    return {
        'note_id': generate_id(row),
        'source': 'weibo',
        'nickname': row.get('博主昵称', ''),
        'content': content_result['original'],
        'content_cleaned': content_result['author_text'],  # 使用清洗后的当前用户发言
        'publish_time': parse_weibo_time(row.get('发布时间')),
        'liked_count': parse_int_safe(row.get('点赞数')),
        'comment_count': parse_int_safe(row.get('评论数')),
        'share_count': parse_int_safe(row.get('转发数')),
        'source_url': row.get('详情链接', ''),
        'source_keyword': keyword,
        'video_url': row.get('视频链接', ''),
        'topics': topics,
        'mentions': content_result['mentions'],
        'amounts': content_result['amounts'],
        'locations': content_result['locations'],
        'forward_chain': content_result['forward_chain'],
    }


def create_table(cur):
    """创建通用表结构"""
    sql = f"""
    CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
        note_id VARCHAR(50) PRIMARY KEY,
        source VARCHAR(20) NOT NULL,
        nickname VARCHAR(100),
        content TEXT,
        content_cleaned TEXT,
        publish_time TIMESTAMP,
        liked_count INTEGER DEFAULT 0,
        comment_count INTEGER DEFAULT 0,
        share_count INTEGER DEFAULT 0,
        source_url TEXT,
        source_keyword VARCHAR(100),
        video_url TEXT,
        topics JSONB,
        mentions JSONB,
        amounts JSONB,
        locations JSONB,
        forward_chain JSONB,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    """
    cur.execute(sql)

    indexes = [
        f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_source ON {TABLE_NAME}(source);",
        f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_nickname ON {TABLE_NAME}(nickname);",
        f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_keyword ON {TABLE_NAME}(source_keyword);",
        f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_topics ON {TABLE_NAME} USING GIN(topics);",
        f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_mentions ON {TABLE_NAME} USING GIN(mentions);",
        f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_publish_time ON {TABLE_NAME}(publish_time);",
    ]
    for idx in indexes:
        cur.execute(idx)

    print(f"表 {TABLE_NAME} 创建完成")


def load_excel_files(dir_path: str, test_file: str = None) -> list:
    """加载所有Excel文件"""
    if test_file:
        files = [os.path.join(dir_path, test_file)]
    else:
        files = glob(os.path.join(dir_path, '*.xlsx'))
    print(f"找到 {len(files)} 个Excel文件")

    records = []
    for file_path in files:
        print(f"  处理: {os.path.basename(file_path)}")
        wb = load_workbook(file_path, data_only=True)
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]
        # 动态查找博主昵称列索引
        nickname_idx = headers.index('博主昵称') if '博主昵称' in headers else 2

        for row_idx in range(2, sheet.max_row + 1):
            row_values = [cell.value for cell in sheet[row_idx]]
            if row_values[nickname_idx]:  # 博主昵称不为空
                row_dict = dict(zip(headers, row_values))
                records.append(row_dict)

        wb.close()

    return records


def import_weibo_data(dir_path: str = None, table_name: str = None):
    """主导入函数"""
    global TABLE_NAME
    conn = None
    cur = None

    # 使用传入的表名或默认表名
    if table_name:
        TABLE_NAME = table_name
        print(f"使用表名: {TABLE_NAME}")

    # 使用传入的目录或默认目录
    excel_dir = dir_path or EXCEL_DIR

    try:
        print("连接数据库...")
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        create_table(cur)
        conn.commit()

        print(f"加载Excel数据...")
        all_records = load_excel_files(excel_dir, TEST_FILE)
        print(f"总记录数: {len(all_records)}")

        # 预处理
        processed = []
        filtered_count = 0
        errors = []

        for row in all_records:
            try:
                result = preprocess_weibo_note(row)
                if result is None:
                    filtered_count += 1  # 短文本过滤
                else:
                    processed.append(result)
            except Exception as e:
                errors.append({'row': row.get('详情链接'), 'error': str(e)})

        print(f"预处理完成: {len(processed)} 条")
        print(f"短文本过滤: {filtered_count} 条")
        print(f"预处理错误: {len(errors)} 条")

        if errors:
            print("\n前5个错误:")
            for err in errors[:5]:
                print(f"  {err}")

        if not processed:
            print("无数据导入")
            return

        # 批量插入
        insert_sql = f"""
        INSERT INTO {TABLE_NAME} (
            note_id, source, nickname, content, content_cleaned, publish_time,
            liked_count, comment_count, share_count, source_url, source_keyword,
            video_url, topics, mentions, amounts, locations, forward_chain
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s::jsonb, %s::jsonb, %s::jsonb, %s::jsonb
        ) ON CONFLICT (note_id) DO UPDATE SET
            content = EXCLUDED.content,
            content_cleaned = EXCLUDED.content_cleaned,
            liked_count = EXCLUDED.liked_count,
            comment_count = EXCLUDED.comment_count,
            share_count = EXCLUDED.share_count,
            topics = EXCLUDED.topics,
            mentions = EXCLUDED.mentions,
            locations = EXCLUDED.locations,
            forward_chain = EXCLUDED.forward_chain
        """

        batch_data = [
            (
                rec['note_id'],
                rec['source'],
                rec['nickname'],
                rec['content'],
                rec['content_cleaned'],
                rec['publish_time'],
                rec['liked_count'],
                rec['comment_count'],
                rec['share_count'],
                rec['source_url'],
                rec['source_keyword'],
                rec['video_url'],
                json.dumps(rec['topics']),
                json.dumps(rec['mentions']),
                json.dumps(rec['amounts']),
                json.dumps(rec['locations']),
                json.dumps(rec['forward_chain']),
            )
            for rec in processed
        ]

        print("导入数据...")
        batch_size = 100
        for i in range(0, len(batch_data), batch_size):
            batch = batch_data[i:i + batch_size]
            execute_batch(cur, insert_sql, batch)
            if (i + batch_size) % 1000 == 0 or i + batch_size >= len(batch_data):
                print(f"  {min(i + batch_size, len(batch_data))}/{len(batch_data)}")
                conn.commit()

        conn.commit()
        print(f"\n导入完成: {len(batch_data)} 条")

        # 统计
        print_stats(cur)

    except Exception as e:
        print(f"错误: {e}")
        if conn:
            conn.rollback()
        raise
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


def print_stats(cur):
    """打印统计"""
    print("\n" + "=" * 50)
    print("数据统计:")
    print("=" * 50)

    cur.execute(f"SELECT COUNT(*) FROM {TABLE_NAME} WHERE source = 'weibo'")
    print(f"  微博记录: {cur.fetchone()[0]}")

    cur.execute(f"""
        SELECT source_keyword, COUNT(*) as cnt
        FROM {TABLE_NAME}
        WHERE source_keyword IS NOT NULL AND source_keyword != ''
        GROUP BY source_keyword
        ORDER BY cnt DESC
        LIMIT 10
    """)
    print("\n  TOP10 关键词:")
    for kw, cnt in cur.fetchall():
        print(f"    {kw}: {cnt}")

    cur.execute(f"""
        SELECT jsonb_array_elements_text(topics) as topic, COUNT(*) as cnt
        FROM {TABLE_NAME}
        WHERE topics IS NOT NULL AND jsonb_array_length(topics) > 0
        GROUP BY topic
        ORDER BY cnt DESC
        LIMIT 10
    """)
    print("\n  TOP10 话题:")
    for topic, cnt in cur.fetchall():
        print(f"    {topic}: {cnt}")

    cur.execute(f"""
        SELECT jsonb_array_elements_text(mentions) as mention, COUNT(*) as cnt
        FROM {TABLE_NAME}
        WHERE mentions IS NOT NULL AND jsonb_array_length(mentions) > 0
        GROUP BY mention
        ORDER BY cnt DESC
        LIMIT 10
    """)
    print("\n  TOP10 @用户:")
    for mention, cnt in cur.fetchall():
        print(f"    {mention}: {cnt}")


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='微博Excel导入PostgreSQL')
    parser.add_argument('--dir', default=None, help='Excel文件目录（默认使用 EXCEL_DIR）')
    parser.add_argument('--table', default=None, help='目标表名（默认使用 social_media_notes，指定新表名可重新导入对比）')
    args = parser.parse_args()
    import_weibo_data(args.dir, args.table)